"""Products — public listing, admin CRUD, image upload, import/export, reviews."""
import io
import os
import time
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..auth import get_current_admin
from ..config import UPLOAD_DIR
from ..database import get_db
from ..models import AdminUser, Brand, Category, Concern, Product, Review, SubCategory
from ..schemas import BrandIn, ProductIn, ProductOut, SubCategoryIn
from ..utils import discount_pct, slugify

router = APIRouter(tags=["products"])


def to_out(p: Product) -> dict:
    return {
        "id": p.id,
        "name": p.name,
        "category_id": p.category_id,
        "subcategory_id": p.subcategory_id or "",
        "brand_id": p.brand_id or "",
        "concerns": p.concerns or [],
        "sku": p.sku or "",
        "price": p.price,
        "mrp": p.mrp,
        "rating": p.rating,
        "reviews_count": p.reviews_count,
        "emoji": p.emoji,
        "badge": p.badge or "",
        "image": p.image or "",
        "gallery": p.gallery or [],
        "video_url": p.video_url or "",
        "description": p.description or "",
        "specifications": p.specifications or {},
        "benefits": p.benefits or [],
        "ingredients": p.ingredients or "",
        "faqs": p.faqs or [],
        "sizes": p.sizes or [],
        "variants": p.variants or [],
        "hsn_code": p.hsn_code or "",
        "gst_rate": p.gst_rate,
        "stock": p.stock,
        "low_stock_threshold": p.low_stock_threshold,
        "weight_grams": p.weight_grams,
        "active": p.active,
        "featured": p.featured,
        "seo_title": p.seo_title or "",
        "seo_description": p.seo_description or "",
        "discount_pct": discount_pct(p.price, p.mrp),
    }


# ───────────────────────────── Public ─────────────────────────────

@router.get("/api/products")
def list_products(
    db: Session = Depends(get_db),
    category: str | None = Query(None),
    subcategory: str | None = Query(None),
    concern: str | None = Query(None),
    brand: str | None = Query(None),
    search: str | None = Query(None),
    featured: bool | None = Query(None),
    include_inactive: bool = Query(False),
):
    q = db.query(Product)
    if not include_inactive:
        q = q.filter(Product.active == True)
    if category:
        q = q.filter(Product.category_id == category)
    if subcategory:
        q = q.filter(Product.subcategory_id == subcategory)
    if brand:
        q = q.filter(Product.brand_id == brand)
    if featured is not None:
        q = q.filter(Product.featured == featured)
    items = q.all()
    out = [to_out(p) for p in items]
    if concern:
        out = [p for p in out if concern in p["concerns"]]
    if search:
        s = search.lower()
        out = [p for p in out if s in (p["name"] + " " + p["description"] + " " + p.get("ingredients", "")).lower()]
    return out


@router.get("/api/products/{product_id}")
def get_product(product_id: str, db: Session = Depends(get_db)):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    return to_out(p)


@router.get("/api/products/{product_id}/reviews")
def product_reviews(product_id: str, db: Session = Depends(get_db)):
    reviews = (
        db.query(Review)
        .filter(Review.product_id == product_id, Review.approved == True)
        .order_by(Review.created_at.desc())
        .all()
    )
    return [
        {
            "id": r.id, "rating": r.rating, "title": r.title, "body": r.body,
            "images": r.images, "verified_purchase": r.verified_purchase,
            "customer_name": r.customer.name if r.customer else "Anonymous",
            "created_at": r.created_at.isoformat(),
        }
        for r in reviews
    ]


# ───────────────────────────── Brands ─────────────────────────────

@router.get("/api/brands")
def list_brands(db: Session = Depends(get_db)):
    return [
        {"id": b.id, "name": b.name, "logo": b.logo, "description": b.description}
        for b in db.query(Brand).order_by(Brand.sort_order).all()
    ]


@router.post("/api/admin/brands")
def upsert_brand(data: BrandIn, db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    bid = data.id or slugify(data.name)
    b = db.query(Brand).filter(Brand.id == bid).first()
    if not b:
        b = Brand(id=bid)
        db.add(b)
    b.name = data.name
    b.logo = data.logo
    b.description = data.description
    b.sort_order = data.sort_order
    db.commit()
    return {"ok": True, "id": b.id}


@router.delete("/api/admin/brands/{bid}")
def delete_brand(bid: str, db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    b = db.query(Brand).filter(Brand.id == bid).first()
    if b:
        db.delete(b)
        db.commit()
    return {"ok": True}


# ───────────────────────────── SubCategories ─────────────────────────────

@router.get("/api/subcategories")
def list_subcategories(db: Session = Depends(get_db), category_id: str | None = Query(None)):
    q = db.query(SubCategory).order_by(SubCategory.sort_order)
    if category_id:
        q = q.filter(SubCategory.category_id == category_id)
    return [
        {"id": s.id, "category_id": s.category_id, "name": s.name, "emoji": s.emoji}
        for s in q.all()
    ]


@router.post("/api/admin/subcategories")
def upsert_subcategory(data: SubCategoryIn, db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    sid = data.id or slugify(data.name)
    s = db.query(SubCategory).filter(SubCategory.id == sid).first()
    if not s:
        s = SubCategory(id=sid, category_id=data.category_id)
        db.add(s)
    s.name = data.name
    s.emoji = data.emoji
    s.sort_order = data.sort_order
    s.category_id = data.category_id
    db.commit()
    return {"ok": True, "id": s.id}


@router.delete("/api/admin/subcategories/{sid}")
def delete_subcategory(sid: str, db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    s = db.query(SubCategory).filter(SubCategory.id == sid).first()
    if s:
        db.delete(s)
        db.commit()
    return {"ok": True}


# ───────────────────────────── Admin CRUD ─────────────────────────────

@router.get("/api/admin/products")
def admin_list_products(
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
    search: str | None = Query(None),
    category: str | None = Query(None),
):
    q = db.query(Product)
    if category:
        q = q.filter(Product.category_id == category)
    if search:
        s = f"%{search}%"
        q = q.filter(Product.name.ilike(s))
    return [to_out(p) for p in q.all()]


@router.post("/api/admin/products", response_model=ProductOut)
def create_product(
    data: ProductIn,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    pid = data.id or slugify(data.name)
    if db.query(Product).filter(Product.id == pid).first():
        pid = f"{pid}-{int(time.time())%10000}"
    p = Product(id=pid)
    for field in [
        "name", "category_id", "subcategory_id", "brand_id", "concerns", "sku",
        "price", "mrp", "rating", "reviews_count", "emoji", "badge", "image",
        "gallery", "video_url", "description", "specifications", "benefits",
        "ingredients", "faqs", "sizes", "variants", "hsn_code", "gst_rate",
        "stock", "low_stock_threshold", "weight_grams", "active", "featured",
        "seo_title", "seo_description",
    ]:
        val = getattr(data, field)
        if field in ("subcategory_id", "brand_id") and val == "":
            val = None
        setattr(p, field, val)
    db.add(p)
    db.commit()
    return to_out(p)


@router.put("/api/admin/products/{product_id}", response_model=ProductOut)
def update_product(
    product_id: str,
    data: ProductIn,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    for field in [
        "name", "category_id", "subcategory_id", "brand_id", "concerns", "sku",
        "price", "mrp", "rating", "reviews_count", "emoji", "badge", "image",
        "gallery", "video_url", "description", "specifications", "benefits",
        "ingredients", "faqs", "sizes", "variants", "hsn_code", "gst_rate",
        "stock", "low_stock_threshold", "weight_grams", "active", "featured",
        "seo_title", "seo_description",
    ]:
        val = getattr(data, field)
        if field in ("subcategory_id", "brand_id") and val == "":
            val = None
        setattr(p, field, val)
    db.commit()
    return to_out(p)


@router.delete("/api/admin/products/{product_id}")
def delete_product(
    product_id: str,
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    p = db.query(Product).filter(Product.id == product_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Product not found")
    db.delete(p)
    db.commit()
    return {"ok": True}


# ───────────────────────────── Admin Reviews ─────────────────────────────

@router.get("/api/admin/reviews")
def admin_list_reviews(
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
    status: str | None = Query(None),
):
    q = db.query(Review).order_by(Review.created_at.desc())
    if status == "pending":
        q = q.filter(Review.approved == False)
    elif status == "approved":
        q = q.filter(Review.approved == True)
    return [
        {
            "id": r.id, "product_id": r.product_id, "customer_id": r.customer_id,
            "rating": r.rating, "title": r.title, "body": r.body,
            "verified_purchase": r.verified_purchase, "approved": r.approved,
            "created_at": r.created_at.isoformat(),
        }
        for r in q.all()
    ]


@router.put("/api/admin/reviews/{review_id}/approve")
def approve_review(review_id: int, db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    r = db.query(Review).filter(Review.id == review_id).first()
    if not r:
        raise HTTPException(status_code=404, detail="Review not found")
    r.approved = True
    # Update product review count
    p = db.query(Product).filter(Product.id == r.product_id).first()
    if p:
        approved_count = db.query(Review).filter(Review.product_id == p.id, Review.approved == True).count()
        p.reviews_count = approved_count
        # Recalculate avg rating
        from sqlalchemy import func
        avg = db.query(func.avg(Review.rating)).filter(Review.product_id == p.id, Review.approved == True).scalar()
        if avg:
            p.rating = round(float(avg), 1)
    db.commit()
    return {"ok": True}


@router.delete("/api/admin/reviews/{review_id}")
def delete_review(review_id: int, db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    r = db.query(Review).filter(Review.id == review_id).first()
    if r:
        db.delete(r)
        db.commit()
    return {"ok": True}


# ───────────────────────────── Image upload ─────────────────────────────

@router.post("/api/admin/upload")
def upload_image(
    file: UploadFile = File(...),
    admin: AdminUser = Depends(get_current_admin),
):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}:
        raise HTTPException(status_code=400, detail="Unsupported image type")
    name = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(UPLOAD_DIR, name)
    with open(path, "wb") as f:
        f.write(file.file.read())
    return {"url": f"/uploads/{name}"}


# ───────────────────────────── Excel / CSV bulk import & export ─────────────────────────────

PRODUCT_COLUMNS = [
    "id", "name", "category_id", "subcategory_id", "brand_id", "concerns",
    "sku", "price", "mrp", "rating", "reviews_count", "emoji", "badge",
    "image", "description", "benefits", "ingredients", "sizes", "hsn_code",
    "gst_rate", "stock", "active",
]


@router.get("/api/admin/products/export.xlsx")
def export_products(db: Session = Depends(get_db), admin: AdminUser = Depends(get_current_admin)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(PRODUCT_COLUMNS)
    for p in db.query(Product).all():
        ws.append([
            p.id, p.name, p.category_id, p.subcategory_id or "", p.brand_id or "",
            ", ".join(p.concerns or []), p.sku or "", p.price, p.mrp,
            p.rating, p.reviews_count, p.emoji, p.badge or "", p.image or "",
            p.description or "", " | ".join(p.benefits or []),
            p.ingredients or "", ", ".join(p.sizes or []),
            p.hsn_code or "", p.gst_rate, p.stock,
            "yes" if p.active else "no",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=primenutra-products.xlsx"},
    )


@router.get("/api/admin/products/template.xlsx")
def template_products(admin: AdminUser = Depends(get_current_admin)):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(PRODUCT_COLUMNS)
    ws.append([
        "tulsi-green-classic", "Tulsi Green Tea Classic", "teas", "", "",
        "immunity, metabolism", "SKU001", 199, 240, 4.7, 1240, "🍵",
        "Bestseller", "", "Soothing blend of holy basil and green tea.",
        "Rich in antioxidants | Supports immunity",
        "Tulsi, Green Tea, Spearmint", "25 Teabags, 50 Teabags",
        "0902", 5, 100, "yes",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=primenutra-products-template.xlsx"},
    )


def _split_list(val, sep):
    if val is None:
        return []
    return [x.strip() for x in str(val).replace("|", sep).split(sep) if x.strip()]


@router.post("/api/admin/products/import")
def import_products(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    admin: AdminUser = Depends(get_current_admin),
):
    fname = (file.filename or "").lower()
    raw = file.file.read()
    rows = []
    if fname.endswith(".csv"):
        import csv
        text = raw.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif fname.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in r):
                continue
            rows.append({header[i]: r[i] for i in range(len(header)) if i < len(r)})
    else:
        raise HTTPException(status_code=400, detail="Upload a .xlsx or .csv file")

    valid_categories = {c.id for c in db.query(Category).all()}
    valid_concerns = {c.id for c in db.query(Concern).all()}
    created, updated, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        try:
            name = (row.get("name") or "").strip()
            if not name:
                continue
            pid = (str(row.get("id")).strip() if row.get("id") else "") or slugify(name)
            cat = (row.get("category_id") or "").strip()
            if cat and cat not in valid_categories:
                errors.append(f"Row {i}: unknown category '{cat}'")
                continue
            concerns = [c for c in _split_list(row.get("concerns"), ",") if not valid_concerns or c in valid_concerns]
            benefits = _split_list(row.get("benefits"), "|") or _split_list(row.get("benefits"), ",")
            sizes = _split_list(row.get("sizes"), ",")

            def num(v, default=0.0):
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return default

            active_raw = str(row.get("active", "yes")).strip().lower()
            active = active_raw in {"yes", "true", "1", "y", ""}

            p = db.query(Product).filter(Product.id == pid).first()
            is_new = p is None
            if is_new:
                p = Product(id=pid)
                db.add(p)
            p.name = name
            p.category_id = cat or p.category_id
            if concerns:
                p.concerns = concerns
            p.sku = (row.get("sku") or p.sku or "")
            p.price = num(row.get("price"), p.price or 0)
            p.mrp = num(row.get("mrp"), p.mrp or 0)
            p.rating = num(row.get("rating"), p.rating or 4.5)
            p.reviews_count = int(num(row.get("reviews_count"), p.reviews_count or 0))
            p.emoji = (row.get("emoji") or p.emoji or "🌿")
            p.badge = (row.get("badge") or "")
            p.image = (row.get("image") or p.image or "")
            p.description = (row.get("description") or p.description or "")
            p.ingredients = (row.get("ingredients") or p.ingredients or "")
            p.hsn_code = (row.get("hsn_code") or p.hsn_code or "")
            p.gst_rate = num(row.get("gst_rate"), p.gst_rate or 18.0)
            if benefits:
                p.benefits = benefits
            if sizes:
                p.sizes = sizes
            p.stock = int(num(row.get("stock"), p.stock or 100))
            p.active = active
            created += 1 if is_new else 0
            updated += 0 if is_new else 1
        except Exception as e:
            errors.append(f"Row {i}: {e}")

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}
